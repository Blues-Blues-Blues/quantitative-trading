"""绩效评估、收益归因与实时流输出模块单元测试（合成 mock 数据，不联网）。

覆盖：
- StreamLogger：JSONL 落盘 / 内存生成器 / 标准字段结构
- PerformanceAnalyzer.analyze：年化收益率、夏普、Calmar、Sortino、回撤、
  平均持仓周期、胜率/盈亏比、日收益偏度/峰度（精确值与端到端）
- 参数敏感度热力图（Optuna study / DataFrame 两种输入）
- AttributionEngine：因子暴露分解归因（盈亏守恒）、IC/Rank IC/IR 双模式
- 20 日人工复盘清单导出（Excel 三 sheet / CSV）
- 四子图 Dashboard 落盘
"""

import os
import sys

import numpy as np
import pandas as pd
import pytest

sys.path.insert(0, os.path.dirname(__file__))
from test_optimizer import TRADE_PARAMS, bull_slice  # noqa: E402

from analytics.attribution import AttributionEngine  # noqa: E402
from analytics.metrics import closed_trades  # noqa: E402
from analytics.performance import PerformanceAnalyzer  # noqa: E402
from analytics.real_time_stream import (  # noqa: E402
    REQUIRED_FIELDS, StreamLogger, to_stream_frame,
)
from optimizer.bayesian_opt import StrategyOptimizer  # noqa: E402
from strategy.signals import Signal  # noqa: E402

_MAPPING = {"600000": "银行"}


@pytest.fixture(scope="module")
def pipeline():
    """跑通端到端回测：返回 (equity_curve, trade_log)。"""
    ds = bull_slice()
    opt = StrategyOptimizer(data=ds, symbol_to_industry=_MAPPING,
                            account_kwargs={"initial_cash": 1e8})
    _, engine = opt.backtest(ds, TRADE_PARAMS)
    log, curve = engine.run()
    return curve, log


# ----------------------------------------------------------------------
# 实时流（JSONL + 生成器）
# ----------------------------------------------------------------------

def _mk_signals():
    return [
        Signal(symbol="600000", timestamp=pd.Timestamp("2024-01-02 10:00"),
               action="HOLD", state="S_noise",
               metrics={"final_ms": -0.2, "global_mod": 0.1, "chain_mod": 0.0,
                        "capital_purity": 0.2, "agent_ms": -0.1}),
        Signal(symbol="600000", timestamp=pd.Timestamp("2024-01-02 10:30"),
               action="BUY", state="S_push",
               metrics={"final_ms": 1.5, "global_mod": 0.7, "chain_mod": 0.3,
                        "capital_purity": 0.55, "agent_ms": 0.5}),
    ]


class TestStreamLogger:

    def test_jsonl_roundtrip(self, tmp_path):
        path = tmp_path / "stream.jsonl"
        with StreamLogger(path) as logger:
            n = logger.log_frame(_mk_signals())
            assert n == 2
        df = StreamLogger.read(path)
        assert len(df) == 2
        for field in REQUIRED_FIELDS:
            assert field in df.columns
        row = df.iloc[1]
        assert row["Action"] == "BUY"
        assert row["State"] == "S_push"
        assert float(row["Final_MS"]) == pytest.approx(1.5)
        assert float(row["Global_Mod"]) == pytest.approx(0.7)
        assert str(row["timestamp"]).startswith("2024-01-02T10:30")

    def test_generator(self):
        rows = list(StreamLogger.generator(_mk_signals()))
        assert len(rows) == 2
        assert rows[0]["symbol"] == "600000"
        assert set(REQUIRED_FIELDS) <= set(rows[0].keys())

    def test_to_stream_frame(self):
        df = to_stream_frame(_mk_signals())
        assert list(df.columns[:len(REQUIRED_FIELDS)]) == REQUIRED_FIELDS
        assert len(df) == 2

    def test_log_without_path_returns_dict(self):
        logger = StreamLogger()
        row = logger.log(_mk_signals()[0])
        assert row["Action"] == "HOLD"
        assert logger.count == 0  # 未落盘


# ----------------------------------------------------------------------
# 绩效指标（精确值 + 端到端）
# ----------------------------------------------------------------------

def _curve(equity: list, days=("2024-01-02", "2024-01-03", "2024-01-04")):
    """每天 4 根 bar；equity 每个元素为当日净值（广播到当日全部 bar）。"""
    idx = pd.DatetimeIndex([])
    for d in days:
        idx = idx.append(pd.date_range(f"{d} 10:00", periods=4, freq="30min"))
    vals = np.repeat(equity, 4)
    assert len(idx) == len(vals)
    return pd.DataFrame({"ts": idx, "total_equity": vals})


class TestPerformanceAnalyzer:

    def test_analyze_end_to_end(self, pipeline):
        curve, log = pipeline
        m = PerformanceAnalyzer.analyze(curve, log)
        assert m["n_trades"] >= 1
        assert 0.0 <= m["win_rate"] <= 1.0
        for key in ("annual_return", "sharpe", "calmar", "sortino",
                    "max_drawdown", "avg_holding_minutes", "profit_loss_ratio",
                    "total_pnl", "daily_skew", "daily_kurtosis"):
            assert key in m

    def test_annual_return_precise(self):
        # 3 个日末值 100 → 110，年化 = (1.1)^(244/3) - 1
        curve = _curve([100.0, 105.0, 110.0])
        assert PerformanceAnalyzer.analyze(curve, pd.DataFrame())["annual_return"] \
            == pytest.approx(1.1 ** (244 / 3) - 1)

    def test_calmar_zero_drawdown_inf(self):
        # 单调上涨 → 回撤 0 → Calmar = +inf（盈利）
        curve = _curve([100.0, 101.0, 103.0])
        assert np.isposinf(PerformanceAnalyzer.analyze(curve, pd.DataFrame())["calmar"])

    def test_sortino_no_downside_inf(self):
        curve = _curve([100.0, 101.0, 103.0])
        assert np.isposinf(PerformanceAnalyzer.analyze(curve, pd.DataFrame())["sortino"])

    def test_avg_holding_period(self, pipeline):
        curve, log = pipeline
        m = PerformanceAnalyzer.analyze(curve, log)
        assert np.isfinite(m["avg_holding_minutes"]) and m["avg_holding_minutes"] >= 0

    def test_analyze_empty_log(self):
        curve = _curve([100.0, 100.0, 100.0])
        m = PerformanceAnalyzer.analyze(curve, pd.DataFrame())
        assert m["n_trades"] == 0
        assert np.isnan(m["win_rate"])

    def test_sensitivity_heatmap_dataframe(self):
        rng = np.random.default_rng(0)
        df = pd.DataFrame({
            "w_ofss": rng.uniform(0.2, 0.6, 60),
            "w_cps": rng.uniform(0.1, 0.5, 60),
            "value": rng.normal(1.0, 0.5, 60),
        })
        import matplotlib.pyplot as plt
        fig, ax = plt.subplots()
        PerformanceAnalyzer.sensitivity_heatmap(df, ax=ax)
        assert ax.get_title() == "Parameter sensitivity (objective mean)"
        plt.close(fig)

    def test_sensitivity_heatmap_insufficient(self):
        import matplotlib.pyplot as plt
        df = pd.DataFrame({"w_ofss": [0.3] * 6, "w_cps": [0.2] * 6,
                           "value": [1.0] * 6})
        fig, ax = plt.subplots()
        PerformanceAnalyzer.sensitivity_heatmap(df, ax=ax)
        assert "insufficient" in ax.texts[0].get_text()
        plt.close(fig)


# ----------------------------------------------------------------------
# 归因：因子暴露分解 + IC
# ----------------------------------------------------------------------

class TestAttributionEngine:

    @staticmethod
    def _signals_for(trades):
        rows = []
        for i, t in enumerate(trades):
            rows.append({"timestamp": t["entry_ts"], "symbol": t["symbol"],
                         "action": "BUY", "state": "S_push",
                         "global_mod": [0.7, 0.1, 0.5][i % 3],
                         "chain_mod": [0.2, 0.8, 0.3][i % 3],
                         "agent_ms": [0.5, 0.4, 0.2][i % 3]})
        return pd.DataFrame(rows)

    def test_attribute_conserves_pnl(self, pipeline):
        curve, log = pipeline
        trades = closed_trades(log)
        sig = self._signals_for(trades)
        trades_df, summary = AttributionEngine.attribute(log, sig)
        assert len(trades_df) == len(trades)
        # 盈亏守恒：逐笔拆解之和 = 归因合计 = 总盈亏
        assert float(trades_df["pnl"].sum()) == pytest.approx(
            float(summary["pnl"].sum()))
        assert float(summary["pnl"].sum()) == pytest.approx(
            sum(t["pnl"] for t in trades))
        # 入场快照可得 → 应有明确主导因子（非 other）
        assert (trades_df["factor"] != "other").all()
        # 每笔 pnl_global + pnl_chain + pnl_agent == pnl
        p = trades_df
        assert np.allclose(p["pnl_global_mod"] + p["pnl_chain_mod"]
                           + p["pnl_agent_ms"], p["pnl"])

    def test_attribute_no_signals_other(self, pipeline):
        curve, log = pipeline
        _, summary = AttributionEngine.attribute(log, pd.DataFrame())
        other = summary[summary["factor"] == "other"].iloc[0]
        assert other["n_trades"] == len(closed_trades(log))

    def test_ic_time_series_mode(self):
        kline, features = _mk_ic_data(n_symbols=1)
        summary, ic_ts = AttributionEngine.compute_ic(
            features, kline, forward=60, window=8)
        assert {"final_ms", "inst_flow"} <= set(summary["factor"])
        row = summary[summary["factor"] == "final_ms"].iloc[0]
        assert np.isfinite(row["rank_ic_mean"])
        assert not ic_ts.empty

    def test_ic_cross_sectional_mode(self):
        kline, features = _mk_ic_data(n_symbols=3)
        summary, ic_ts = AttributionEngine.compute_ic(
            features, kline, forward="1D")
        assert not ic_ts.empty
        row = summary[summary["factor"] == "final_ms"].iloc[0]
        assert np.isfinite(row["rank_ic_mean"])
        assert np.isfinite(row["ic_mean"]) or np.isfinite(row["ic_ir"])

    def test_plot_attribution(self, pipeline, tmp_path):
        curve, log = pipeline
        _, summary = AttributionEngine.attribute(
            log, self._signals_for(closed_trades(log)))
        path = AttributionEngine.plot_attribution(
            summary, str(tmp_path / "attr.png"))
        assert os.path.exists(path)

    def test_plot_ic_heatmap(self, tmp_path):
        kline, features = _mk_ic_data(n_symbols=3)
        _, ic_ts = AttributionEngine.compute_ic(features, kline, forward="1D")
        path = AttributionEngine.plot_ic_heatmap(
            ic_ts, str(tmp_path / "ic.png"), buckets=4)
        assert os.path.exists(path)


# ----------------------------------------------------------------------
# 20 日复盘清单 + Dashboard
# ----------------------------------------------------------------------

class TestReviewExport:

    def test_export_xlsx(self, pipeline, tmp_path):
        curve, log = pipeline
        path = PerformanceAnalyzer.export_review_slices(
            bull_slice(), log, days=20, path=str(tmp_path / "review.xlsx"))
        assert os.path.exists(path)
        import openpyxl
        wb = openpyxl.load_workbook(path)
        assert set(wb.sheetnames) == {"summary", "daily_slices", "tick_flows"}
        ws = wb["summary"]
        assert ws.max_row >= 2  # 表头 + 至少一笔交易
        assert ws.cell(1, 2).value == "symbol"

    def test_export_csv(self, pipeline, tmp_path):
        curve, log = pipeline
        out = PerformanceAnalyzer.export_review_slices(
            bull_slice(), log, fmt="csv", path=str(tmp_path / "review_csv"))
        for name in ("summary", "daily_slices", "tick_flows"):
            assert os.path.exists(os.path.join(out, f"{name}.csv"))

    def test_plot_report(self, pipeline, tmp_path):
        curve, log = pipeline
        trades = closed_trades(log)
        _, summary = AttributionEngine.attribute(
            log, TestAttributionEngine._signals_for(trades))
        import matplotlib.pyplot as plt
        fig, ax = plt.subplots()
        PerformanceAnalyzer.sensitivity_heatmap(
            pd.DataFrame({"w_ofss": np.linspace(0.2, 0.6, 20),
                          "w_cps": np.linspace(0.1, 0.5, 20),
                          "value": np.sin(np.linspace(0, 6, 20))}),
            ax=ax)
        plt.close(fig)
        path = PerformanceAnalyzer.plot_report(
            curve, log, attribution_summary=summary,
            study=pd.DataFrame({"w_ofss": np.linspace(0.2, 0.6, 12),
                                "w_cps": np.linspace(0.1, 0.5, 12),
                                "value": np.linspace(0, 2, 12)}),
            path=str(tmp_path / "dashboard.png"))
        assert os.path.exists(path)


# ----------------------------------------------------------------------
# IC mock 数据
# ----------------------------------------------------------------------

def _mk_ic_data(n_symbols=1, days=4):
    """多/单标的 30 分钟 K 线 + 因子长表（因子与未来收益正相关）。"""
    dates = pd.date_range("2024-01-02", periods=days, freq="B")
    parts = []
    for d in dates:
        parts.append(pd.date_range(f"{d:%Y-%m-%d} 09:30", f"{d:%Y-%m-%d} 11:30", freq="30min"))
        parts.append(pd.date_range(f"{d:%Y-%m-%d} 13:00", f"{d:%Y-%m-%d} 15:00", freq="30min"))
    axis = pd.DatetimeIndex(np.concatenate([p.values for p in parts])).sort_values()

    krows, frows = [], []
    for i, t in enumerate(axis):
        for s in range(n_symbols):
            sym = f"60000{s}"
            c = 10.0 + 0.01 * i + 0.001 * s * i
            krows.append({"ts": t, "symbol": sym, "close": c})
            factor = 0.05 * i + 0.5 * s      # 单调上升 → 与未来收益正相关
            frows.append({"ts": t, "symbol": sym, "final_ms": factor,
                          "inst_flow": factor * 0.5})
    kline = pd.DataFrame(krows).set_index("ts")
    features = pd.DataFrame(frows).set_index("ts")
    return kline, features
